from __future__ import annotations
from datetime import datetime

import os
import base64
import hashlib
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook


HEADERS = ["ФИО", "telegram", "жалоба", "рекомендуемый врач", "дата обращения"]


def _try_load_fernet():
    try:
        from cryptography.fernet import Fernet
        return Fernet, True
    except Exception:
        return None, False


def _load_or_create_key(key_path: str) -> bytes:
    if os.path.exists(key_path):
        with open(key_path, "rb") as f:
            return f.read().strip()

    Fernet, ok = _try_load_fernet()
    if not ok:
        raw = os.urandom(32)
        with open(key_path, "wb") as f:
            f.write(raw)
        return raw

    key = Fernet.generate_key()
    with open(key_path, "wb") as f:
        f.write(key)
    return key


def _fallback_encrypt(plaintext: str, key: bytes) -> str:
    data = plaintext.encode("utf-8")
    out = bytearray(len(data))

    counter = 0
    offset = 0
    while offset < len(data):
        block = hashlib.sha256(key + counter.to_bytes(8, "big")).digest()
        for b in block:
            if offset >= len(data):
                break
            out[offset] = data[offset] ^ b
            offset += 1
        counter += 1

    return base64.urlsafe_b64encode(bytes(out)).decode("ascii")


def _encrypt_value(value: str, key: bytes) -> str:
    Fernet, ok = _try_load_fernet()
    if ok:
        f = Fernet(key)
        return f.encrypt(value.encode("utf-8")).decode("ascii")
    return _fallback_encrypt(value, key)


@dataclass
class SecureExcelLogger:
    excel_path: str = "cases.xlsx"
    key_path: str = "secret.key"

    def __post_init__(self):
        self._key = _load_or_create_key(self.key_path)

    def _open_ws(self):
        if os.path.exists(self.excel_path):
            wb = load_workbook(self.excel_path)
            ws = wb.active
            if ws.max_row < 1:
                ws.append(HEADERS)
            else:
                current_headers = [ws.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)]
                if current_headers != HEADERS:
                    for i, header in enumerate(HEADERS, start=1):
                        ws.cell(row=1, column=i).value = header
            return wb, ws

        wb = Workbook()
        ws = wb.active
        ws.title = "Cases"
        ws.append(HEADERS)
        wb.save(self.excel_path)
        return wb, ws

    def create_case(self, fio: str, telegram: str) -> int:
        wb, ws = self._open_ws()

        enc_fio = _encrypt_value(fio, self._key)
        enc_tg = _encrypt_value(telegram, self._key)
        enc_empty = _encrypt_value("", self._key)

        current_dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        enc_dt = _encrypt_value(current_dt, self._key)

        ws.append([enc_fio, enc_tg, enc_empty, enc_empty, enc_dt])
        row_idx = ws.max_row

        wb.save(self.excel_path)
        return row_idx

    def update_case(self, row_idx: int, complaint: str | None = None, doctor: str | None = None) -> None:
        wb, ws = self._open_ws()

        if row_idx < 2 or row_idx > ws.max_row:
            raise ValueError(f"Некорректный row_idx={row_idx}")

        if complaint is not None:
            ws.cell(row=row_idx, column=3).value = _encrypt_value(complaint, self._key)

        if doctor is not None:
            ws.cell(row=row_idx, column=4).value = _encrypt_value(doctor, self._key)

        wb.save(self.excel_path)