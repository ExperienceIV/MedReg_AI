from __future__ import annotations

import base64
import hashlib
import os
from typing import Optional

from openpyxl import load_workbook


def try_load_fernet():
    try:
        from cryptography.fernet import Fernet
        return Fernet, True
    except Exception:
        return None, False


def load_key(key_path: str) -> bytes:
    with open(key_path, "rb") as f:
        return f.read().strip()


def fallback_xor_stream(data: bytes, key: bytes) -> bytes:
    out = bytearray(len(data))
    counter = 0
    offset = 0
    while offset < len(data):
        block = hashlib.sha256(key + counter.to_bytes(8, "big")).digest()
        for b in block:
            if offset >= len(data):
                break
            out[offset] = data[offset] ^ b
            offset += 1
        counter += 1
    return bytes(out)


def decrypt_value(token: Optional[str], key: bytes) -> str:
    if token is None:
        return ""
    token = str(token).strip()
    if token == "":
        return ""

    Fernet, ok = try_load_fernet()
    if ok:
        try:
            f = Fernet(key)
            return f.decrypt(token.encode("ascii")).decode("utf-8")
        except Exception:
            pass

    try:
        raw = base64.urlsafe_b64decode(token.encode("ascii"))
        plain = fallback_xor_stream(raw, key).decode("utf-8", errors="replace")
        return plain
    except Exception:
        return "<не удалось расшифровать>"


def main():
    excel_path = "cases.xlsx"
    key_path = "secret.key"

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Не найден файл {excel_path}")
    if not os.path.exists(key_path):
        raise FileNotFoundError(f"Не найден файл {key_path}")

    key = load_key(key_path)

    wb = load_workbook(excel_path)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        fio = decrypt_value(ws.cell(row=row, column=1).value, key)
        tg = decrypt_value(ws.cell(row=row, column=2).value, key)
        complaint = decrypt_value(ws.cell(row=row, column=3).value, key)
        doctor = decrypt_value(ws.cell(row=row, column=4).value, key)
        request_date = decrypt_value(ws.cell(row=row, column=5).value, key)
        
        request_date = ""
        if ws.max_column >= 5:
            request_date = decrypt_value(ws.cell(row=row, column=5).value, key)

        print(f"\n--- Запись #{row-1} (строка Excel {row}) ---")
        print("ФИО:", fio)
        print("Telegram:", tg)
        print("Жалоба:", complaint)
        print("Рекомендуемый врач:", doctor)
        print("Дата обращения:", request_date)


if __name__ == "__main__":
    main()
